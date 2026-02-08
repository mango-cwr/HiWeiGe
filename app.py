# -*- coding: utf-8 -*-
"""
套餐详情分析 Web 服务：上传套餐文件（Excel）并进行分析。
规则：忽略表格前 10 行与后 7 行；按「套餐」归类——套餐为 A 列中字号 16、加粗的单元格。
"""
import re
import xml.etree.ElementTree as ET
from pathlib import Path

from flask import Flask, render_template, jsonify, request
import pandas as pd
import openpyxl
import time

app = Flask(__name__)
# 当前项目目录即工作目录
BASE_DIR = Path(__file__).resolve().parent
EXCEL_EXTENSIONS = ('.xlsx', '.xls')


def _cell_value(cell):
    """取单元格显示值，兼容空与数字。"""
    if cell.value is None:
        return ''
    return str(cell.value).strip()


def _is_package_name_cell(cell):
    """判断是否为「套餐名」单元格：A 列，字号 16，加粗。"""
    if cell is None:
        return False
    font = cell.font
    if font is None:
        return False
    size = getattr(font, 'size', None)
    bold = getattr(font, 'bold', None)
    return (size is not None and int(size) == 16) and bold is True


def _is_other_product_cell(cell):
    """判断是否为「其他」产品名单元格：A 列，字号 14（接入号形式）。"""
    if cell is None:
        return False
    font = cell.font
    if font is None:
        return False
    size = getattr(font, 'size', None)
    return size is not None and int(size) == 14


def _cell_a_equals_total(cell):
    """判断 A 列单元格内容是否为「合计」（套餐结束标记）。"""
    if cell is None or cell.value is None:
        return False
    return str(cell.value).strip() == '合计'


# 用于识别「金额」列的表头关键词（任一匹配即可）
_AMOUNT_HEADER_KEYWORDS = ('金额', '合计', '费用', '总价', '元)', '（元）', '应收', '应付', '实收', '实付', '小计', '总计')


def _find_amount_column_index(first_row):
    """根据第一行（表头）找金额列索引，找不到返回 None。"""
    if not first_row:
        return None
    for i, cell in enumerate(first_row):
        s = (cell or '').strip()
        if not s:
            continue
        for kw in _AMOUNT_HEADER_KEYWORDS:
            if kw in s:
                return i
    return None


def _parse_number(s):
    """将单元格字符串转为数字，无法解析返回 None。"""
    if s is None:
        return None
    s = str(s).strip().replace(',', '').replace('，', '')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _sum_amount_in_group(rows, amount_col_idx):
    """
    计算一个套餐块的合计金额。

    优先按「实际消费」的小计/合计行汇总：
    - 先在块中寻找包含「原价/减免/实际消费」等表头的那一行；
    - 确定「原价」「减免」「实际消费」所在列；
    - 在该表头下方，只汇总 A 列为「小计」的行；
      - 若没有「小计」行，则退而汇总 A 列为「合计」的行；
    - 金额优先取「实际消费」列；若没有，则用「原价 + 减免」。

    若未能识别上述结构，则退回到旧逻辑：
    - 使用给定的金额列（或最后一列）从第 2 行开始简单求和。
    """
    if not rows:
        return 0.0

    # -------- 优先：识别「原价 / 减免 / 实际消费」结构，并按小计/合计汇总 --------
    header_idx = None
    for i, row in enumerate(rows):
        for cell in row:
            s = str(cell).strip()
            if any(key in s for key in ('原价', '减免', '优惠', '实际消费', '实收', '实付')):
                header_idx = i
                break
        if header_idx is not None:
            break

    if header_idx is not None:
        header = rows[header_idx]
        price_col = None
        discount_col = None
        actual_col = None
        for idx, cell in enumerate(header):
            s = str(cell).strip()
            if '原价' in s:
                price_col = idx
            if '减免' in s or '优惠' in s:
                discount_col = idx
            if '实际消费' in s or '实收' in s or '实付' in s:
                actual_col = idx

        # 只在识别到至少一种金额相关列时才按新规则处理
        if actual_col is not None or price_col is not None or discount_col is not None:
            subtotal_rows = []
            total_rows = []
            for i in range(header_idx + 1, len(rows)):
                first_cell = (rows[i][0] if rows[i] else '').strip()
                if '小计' in first_cell:
                    subtotal_rows.append(i)
                elif first_cell == '合计' or first_cell.startswith('合计'):
                    total_rows.append(i)

            target_indices = subtotal_rows or total_rows

            if target_indices:
                total = 0.0
                for i in target_indices:
                    row = rows[i]
                    value = None
                    # 1) 优先使用「实际消费」列
                    if actual_col is not None and actual_col < len(row):
                        value = _parse_number(row[actual_col])
                    # 2) 没有实际消费列，则尝试「原价 + 减免」
                    if value is None:
                        price = _parse_number(row[price_col]) if price_col is not None and price_col < len(row) else None
                        discount = _parse_number(row[discount_col]) if discount_col is not None and discount_col < len(row) else None
                        if price is not None or discount is not None:
                            value = (price or 0.0) + (discount or 0.0)
                    if value is not None:
                        total += value
                return round(total, 2)

    # -------- 兜底：维持旧有的「从第 2 行起简单求和」逻辑 --------
    total = 0.0
    ncols = max(len(r) for r in rows) if rows else 0
    col = amount_col_idx if amount_col_idx is not None and 0 <= amount_col_idx < ncols else (ncols - 1 if ncols else 0)
    # 从第 1 行开始累加（第 0 行可能是表头或套餐名）
    for i in range(1, len(rows)):
        row = rows[i]
        if col < len(row):
            v = _parse_number(row[col])
            if v is not None:
                total += v
    return round(total, 2)


# 其他套餐中：每个「产品名称：接入号」下方第一个「合计」行，其右边第 4 列（E 列，0-based 索引 4）为消费金额
_OTHER_TOTAL_ROW_RIGHT_OFFSET = 4


def _sum_other_package_total_below_product(rows_data, other_indices):
    """
    对「其他套餐」中每个产品名行（other_indices），找其下方第一个 A 列为「合计」的行，
    取该行「合计」右边第 4 个单元格（E 列）的值，求和并返回。
    """
    total = 0.0
    col_idx = _OTHER_TOTAL_ROW_RIGHT_OFFSET  # D 列
    n = len(rows_data)
    for idx in other_indices:
        for j in range(idx + 1, n):
            if (rows_data[j][0] if rows_data[j] else '').strip() == '合计':
                if col_idx < len(rows_data[j]):
                    v = _parse_number(rows_data[j][col_idx])
                    if v is not None:
                        total += v
                break
    return round(total, 2)


def read_excel_grouped_by_package(path, sheet_name=None):
    """
    套餐归类逻辑：
    - 有效区：忽略前 10 行、后 7 行。
    - 套餐名：A 列，16 号字加粗。套餐从该行开始，到 A 列内容为「合计」的那一行结束（含该行）。
    - 其他套餐：A 列为 14 号字体的单元格（产品名/接入号形式），且不在任一「套餐名→合计」块内的行，归为「其他套餐」。
    """
    path = BASE_DIR / path
    if not path.is_file() or path.suffix.lower() != '.xlsx':
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        sheet_names = wb.sheetnames
        if sheet_name is not None:
            if sheet_name not in sheet_names:
                sheet_names = [sheet_names[0]] if sheet_names else []
            else:
                sheet_names = [sheet_name]
        result_sheets = []
        for sname in sheet_names:
            ws = wb[sname]
            max_row = ws.max_row
            max_col = ws.max_column
            start_row = 11
            end_row = max_row - 7
            if end_row < start_row:
                result_sheets.append({'name': sname, 'groups': []})
                continue
            rows_data = []
            for r in range(start_row, end_row + 1):
                row_list = []
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    row_list.append(_cell_value(cell))
                rows_data.append(row_list)
            # 带格式的工作表：判断 A 列 16 号加粗、14 号、以及内容「合计」
            wb_font = openpyxl.load_workbook(path, data_only=False, read_only=False)
            ws_font = wb_font[sname]
            n = len(rows_data)
            is_package_name = [False] * n  # 该行 A 列是否为 16 号加粗（套餐名）
            is_other_14 = [False] * n     # 该行 A 列是否为 14 号（其他/接入号）
            is_total_row = [False] * n    # 该行 A 列内容是否为「合计」
            for i, r in enumerate(range(start_row, end_row + 1)):
                cell_a = ws_font.cell(row=r, column=1)
                is_package_name[i] = _is_package_name_cell(cell_a)
                is_other_14[i] = _is_other_product_cell(cell_a)
                is_total_row[i] = _cell_a_equals_total(cell_a)
            wb_font.close()
            # 1) 套餐块：从每个「套餐名」行到下一个 A 列「合计」行（含）
            package_blocks = []  # [(start_idx, end_idx), ...]
            i = 0
            while i < n:
                if is_package_name[i]:
                    start_i = i
                    end_i = start_i
                    for j in range(i + 1, n):
                        if is_total_row[j]:
                            end_i = j
                            break
                        if is_package_name[j]:
                            # 下一个套餐开始前没遇到「合计」，则当前套餐到 j-1 止
                            end_i = j - 1
                            break
                    else:
                        end_i = n - 1
                    package_blocks.append((start_i, end_i))
                    i = end_i + 1
                else:
                    i += 1
            in_package = set()
            for (a, b) in package_blocks:
                for idx in range(a, b + 1):
                    in_package.add(idx)
            # 2) 其他套餐：A 列 14 号且不在任一套餐块内的行
            other_indices = [i for i in range(n) if is_other_14[i] and i not in in_package]
            groups = []
            for (start_i, end_i) in package_blocks:
                group_rows = rows_data[start_i:end_i + 1]
                package_name = (rows_data[start_i][0] if rows_data[start_i] else '').strip()
                amount_col = _find_amount_column_index(group_rows[0] if group_rows else None)
                total_amount = _sum_amount_in_group(group_rows, amount_col)
                groups.append({
                    'packageName': package_name or '未命名套餐',
                    'rows': group_rows,
                    'totalAmount': total_amount,
                    'amountCol': amount_col,
                })
            if other_indices:
                # 其他套餐：每个接入号到其下方第一个「合计」行（含）整块保留，便于前端展示明细并按合计行实际消费显示金额
                other_rows = []
                for idx in other_indices:
                    for j in range(idx + 1, n):
                        if (rows_data[j][0] if rows_data[j] else '').strip() == '合计':
                            other_rows.extend(rows_data[idx : j + 1])
                            break
                    else:
                        other_rows.append(rows_data[idx])
                amount_col = _find_amount_column_index(other_rows[0] if other_rows else None)
                # other_rows 已含每个接入号到「合计」的整块，_sum_amount_in_group 会汇总其中小计/合计行，不再重复加 _sum_other_package_total_below_product 避免双倍
                total_amount = _sum_amount_in_group(other_rows, amount_col)
                total_amount = round(total_amount, 2)
                groups.append({
                    'packageName': '其他套餐',
                    'rows': other_rows,
                    'totalAmount': total_amount,
                    'amountCol': amount_col,
                })
            if not groups:
                amount_col = _find_amount_column_index(rows_data[0] if rows_data else None)
                total_amount = _sum_amount_in_group(rows_data, amount_col)
                groups.append({
                    'packageName': '（未识别到套餐，显示全部有效行）',
                    'rows': rows_data,
                    'totalAmount': total_amount,
                })
            result_sheets.append({'name': sname, 'groups': groups})
        sheet_names = list(wb.sheetnames)
        wb.close()
        return {'sheetNames': sheet_names, 'sheets': result_sheets}
    except Exception as e:
        return {'error': str(e)}


def read_excel_sheets(path):
    """读取套餐文件所有工作表，返回各表的数据与基本信息（应用前10行/后7行忽略，按套餐归类）。"""
    path = BASE_DIR / path
    if not path.is_file() or path.suffix.lower() not in EXCEL_EXTENSIONS:
        return None
    if path.suffix.lower() == '.xlsx':
        result = read_excel_grouped_by_package(path, sheet_name=None)
        if result is None:
            return None
        if 'error' in result:
            return result
        # 前端兼容：把 groups 展平为带 packageName 的 data，并保留 totalRows 等
        sheets = []
        for sh in result['sheets']:
            all_rows = []
            for g in sh['groups']:
                all_rows.extend(g['rows'])
            sheets.append({
                'name': sh['name'],
                'rows': len(all_rows),
                'cols': max(len(r) for r in all_rows) if all_rows else 0,
                'numericColumnCount': 0,
                'data': all_rows[:500],
                'totalRows': len(all_rows),
                'groups': sh['groups'],
            })
        return {'sheets': sheets, 'sheetNames': result['sheetNames']}
    try:
        xl = pd.ExcelFile(path, engine=None)
        sheets = []
        for name in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=name, header=None)
            raw = df.fillna('').astype(str)
            rows = raw.values.tolist()
            n_rows, n_cols = df.shape
            sheets.append({
                'name': name,
                'rows': n_rows,
                'cols': n_cols,
                'numericColumnCount': 0,
                'data': rows[:500],
                'totalRows': n_rows,
                'groups': [],
            })
        return {'sheets': sheets, 'sheetNames': xl.sheet_names}
    except Exception as e:
        return {'error': str(e)}


@app.route('/api/upload', methods=['POST'])
def api_upload():
    """
    上传单个套餐文件并直接返回分析结果。

    - 接收表单字段名为 "file" 的文件；
    - 临时保存到 BASE_DIR/uploads 目录；
    - 调用现有的 read_excel_sheets / analyze_sheet_data 进行分析；
    - 返回结构：{ originalName, storedPath, read, analyze }。
    """
    if 'file' not in request.files:
        return jsonify({'error': '缺少文件字段 file'}), 400

    f = request.files['file']
    if f.filename is None or f.filename.strip() == '':
        return jsonify({'error': '文件名为空'}), 400

    suffix = Path(f.filename).suffix.lower()
    if suffix not in EXCEL_EXTENSIONS:
        return jsonify({'error': '仅支持套餐文件（.xls, .xlsx）'}), 400

    uploads_dir = BASE_DIR / 'uploads'
    try:
        uploads_dir.mkdir(exist_ok=True)
    except Exception:
        # 目录创建失败时直接报错
        return jsonify({'error': '无法创建上传目录'}), 500

    ts = int(time.time() * 1000)
    stored_rel_path = Path('uploads') / f'upload_{ts}{suffix}'
    stored_abs_path = BASE_DIR / stored_rel_path

    try:
        f.save(str(stored_abs_path))

        read_result = read_excel_sheets(str(stored_rel_path))
        if read_result is None:
            return jsonify({'error': '文件读取失败'}), 400
        if 'error' in read_result:
            return jsonify({'error': read_result.get('error', '文件解析出错')}), 400

        analyze_result = analyze_sheet_data(str(stored_rel_path))
        if analyze_result is None:
            analyze_result = {}

        return jsonify({
            'originalName': f.filename,
            'storedPath': str(stored_rel_path),
            'read': read_result,
            'analyze': analyze_result,
        })
    finally:
        # 简单清理：尝试删除临时文件，失败则忽略
        try:
            if stored_abs_path.is_file():
                stored_abs_path.unlink()
        except Exception:
            pass


def _parse_billing_cycle_month(cell_value):
    """
    从账单周期单元格解析出月份字符串，如 2024-07。
    格式示例：[20240701]2024-07-01:2024-07-31
    """
    if cell_value is None or pd.isna(cell_value):
        return None
    s = str(cell_value).strip()
    if not s:
        return None
    # 匹配 [20240701]2024-07-01: 或 2024-07-01: 部分，取日期前 7 位为 2024-07
    m = re.search(r'(\d{4}-\d{2})-\d{2}', s)
    if m:
        return m.group(1)
    # 备选：仅 [20240701] 形式，取 20240701 转成 2024-07
    m = re.search(r'\[(\d{4})(\d{2})\d{2}\]', s)
    if m:
        return f'{m.group(1)}-{m.group(2)}'
    return None


# Excel 2003 XML (SpreadsheetML) 命名空间
_XML_NS = 'urn:schemas-microsoft-com:office:spreadsheet'


def _read_excel_2003_xml(path):
    """
    解析 Excel 2003 XML (SpreadsheetML) 格式，返回第一张表的 DataFrame，失败返回 None。
    结构：Workbook -> Worksheet -> Table -> Row -> Cell (ss:Index 可选) -> Data
    """
    try:
        tree = ET.parse(path)
        root = tree.getroot()
        # 可能带命名空间：{urn:...}Worksheet 或 ss:Worksheet
        def local_tag(tag):
            if tag and '}' in tag:
                return tag.split('}', 1)[1]
            return tag or ''
        # 找第一个 Worksheet -> Table
        table = None
        for w in root:
            if local_tag(w.tag) == 'Worksheet':
                for t in w:
                    if local_tag(t.tag) == 'Table':
                        table = t
                        break
                break
        if table is None:
            return None
        rows_data = []
        for row_elem in table:
            if local_tag(row_elem.tag) != 'Row':
                continue
            # 收集该行所有 Cell，Cell 可有 ss:Index (1-based)
            idx_attr = '{' + _XML_NS + '}Index'
            cells = []
            for cell in row_elem:
                if local_tag(cell.tag) != 'Cell':
                    continue
                idx = cell.get(idx_attr) or cell.get('Index')
                col_idx = int(idx) - 1 if idx else len(cells)
                value = ''
                for data in cell:
                    if local_tag(data.tag) == 'Data':
                        if data.text:
                            value = data.text.strip()
                        break
                # 若指定了 Index，前面可能缺列，补空
                while len(cells) < col_idx:
                    cells.append('')
                cells.append(value)
            if cells:
                rows_data.append(cells)
        if not rows_data:
            return None
        # 第一行作为表头，列数取最大
        max_cols = max(len(r) for r in rows_data)
        for r in rows_data:
            while len(r) < max_cols:
                r.append('')
        df = pd.DataFrame(rows_data[1:], columns=rows_data[0] if rows_data else [])
        return df
    except Exception:
        return None


def _excel_engine(path):
    """
    根据文件实际内容选择 pandas 读 Excel 的 engine，避免扩展名 .xls 但内容为 XML/xlsx 时报错。
    - 文件头为 PK：xlsx（ZIP），用 openpyxl
    - 文件头为 D0 CF 11 E0：二进制 .xls（OLE2），用 xlrd
    - 文件头为 <?xml：Excel 2003 XML 等，用 openpyxl 尝试（若失败需用户另存为 .xlsx）
    """
    p = path if isinstance(path, Path) else Path(path)
    if not p.is_file():
        return 'openpyxl'
    try:
        with open(p, 'rb') as f:
            head = f.read(8)
    except Exception:
        head = b''
    if head.startswith(b'PK'):
        return 'openpyxl'
    if head[:4] == b'\xd0\xcf\x11\xe0':
        return 'xlrd'
    if head.startswith(b'<?xml') or head.startswith(b'\xef\xbb\xbf<?xml'):
        return 'openpyxl'
    # 按扩展名兜底
    suffix = p.suffix.lower()
    return 'xlrd' if suffix == '.xls' else 'openpyxl'


def parse_monthly_bill_for_diff(path):
    """
    读取 Excel，按账单周期列、号码列、账单费用列解析，用于月度差异对比。
    返回：{ 'months': ['2024-07', '2024-08', ...], 'byMonth': { '2024-07': [ { 'number', 'fee' }, ... ], ... } }
    列名匹配：账单周期/周期/账期；号码/接入号/手机号；账单费用/费用/实际消费 等。
    """
    path = BASE_DIR / path
    if not path.is_file():
        return None
    try:
        # 若文件头为 Excel 2003 XML，优先用专用解析器，避免 read_excel 报错
        try:
            with open(path, 'rb') as f:
                head = f.read(8)
        except Exception:
            head = b''
        if head.startswith(b'<?xml') or head.startswith(b'\xef\xbb\xbf<?xml'):
            df = _read_excel_2003_xml(path)
            if df is not None:
                pass  # 使用 df 继续下方逻辑
            else:
                df = None  # 交给下方 engine 逻辑
        else:
            df = None
        if df is None:
            engine = _excel_engine(path)
            try:
                df = pd.read_excel(path, sheet_name=0, header=0, engine=engine)
            except Exception as e1:
                err_msg = str(e1)
                # 扩展名 .xls 但内容为 XML 时 xlrd 报 BOF；或 openpyxl 报非 zip
                if 'BOF' in err_msg or '<?xml' in err_msg or 'not a zip' in err_msg.lower() or 'zip' in err_msg.lower():
                    other = 'openpyxl' if engine == 'xlrd' else 'xlrd'
                    try:
                        df = pd.read_excel(path, sheet_name=0, header=0, engine=other)
                    except Exception:
                        df = _read_excel_2003_xml(path)
                        if df is None:
                            return {'error': '文件格式无法识别（可能为 Excel 2003 XML）。请用 Excel 另存为 .xlsx 后重新上传。'}
                else:
                    raise
        if df.empty or len(df.columns) == 0:
            return {'error': '表格为空或无表头'}
        cols = [str(c).strip() for c in df.columns]
        # 找列索引
        cycle_col = None
        number_col = None
        fee_col = None
        cycle_kw = ('账单周期', '周期', '账期', '账务周期', '计费周期')
        number_kw = ('号码', '接入号', '手机号', '产品', '客户', '账户')
        fee_kw = ('账单费用', '费用', '实际消费', '消费', '金额', '合计')
        for i, c in enumerate(cols):
            if cycle_col is None and any(k in c for k in cycle_kw):
                cycle_col = i
            if number_col is None and any(k in c for k in number_kw):
                number_col = i
            if fee_col is None and any(k in c for k in fee_kw):
                fee_col = i
        if cycle_col is None:
            return {'error': '未找到账单周期列（需包含“周期”或“账期”等）'}
        if number_col is None:
            return {'error': '未找到号码列（需包含“号码”或“接入号”等）'}
        if fee_col is None:
            return {'error': '未找到账单费用列（需包含“费用”或“金额”等）'}
        by_month = {}
        for _, row in df.iterrows():
            cycle_val = row.iloc[cycle_col] if cycle_col < len(row) else None
            month = _parse_billing_cycle_month(cycle_val)
            if not month:
                continue
            num_val = row.iloc[number_col] if number_col < len(row) else None
            number = '' if pd.isna(num_val) else str(num_val).strip()
            if not number:
                continue
            fee_val = row.iloc[fee_col] if fee_col < len(row) else None
            fee = _parse_number(fee_val) if fee_val is not None else None
            if fee is None:
                fee = 0.0
            if month not in by_month:
                by_month[month] = []
            # 同一月同一号码可能多行，先按 (月, 号码) 聚合为一条（费用相加）
            found = False
            for item in by_month[month]:
                if item.get('number') == number:
                    item['fee'] = (item.get('fee') or 0) + fee
                    found = True
                    break
            if not found:
                by_month[month].append({'number': number, 'fee': round(fee, 2)})
        months = sorted(by_month.keys())
        if not months:
            return {'error': '未能从账单周期列解析出有效月份'}
        return {'months': months, 'byMonth': by_month}
    except Exception as e:
        return {'error': str(e)}


@app.route('/api/upload_monthly_diff', methods=['POST'])
def api_upload_monthly_diff():
    """
    上传 Excel 用于「月度账单差异对比」。
    解析账单周期列（格式如 [20240701]2024-07-01:2024-07-31）、号码列、账单费用列，
    返回可选的月份列表及按月份汇总的号码与费用，供前端选择两月对比。
    """
    if 'file' not in request.files:
        return jsonify({'error': '缺少文件字段 file'}), 400
    f = request.files['file']
    if f.filename is None or f.filename.strip() == '':
        return jsonify({'error': '文件名为空'}), 400
    suffix = Path(f.filename).suffix.lower()
    if suffix not in EXCEL_EXTENSIONS:
        return jsonify({'error': '仅支持 Excel 文件（.xls, .xlsx）'}), 400
    uploads_dir = BASE_DIR / 'uploads'
    try:
        uploads_dir.mkdir(exist_ok=True)
    except Exception:
        return jsonify({'error': '无法创建上传目录'}), 500
    ts = int(time.time() * 1000)
    stored_rel_path = Path('uploads') / f'monthly_{ts}{suffix}'
    stored_abs_path = BASE_DIR / stored_rel_path
    try:
        f.save(str(stored_abs_path))
        result = parse_monthly_bill_for_diff(str(stored_rel_path))
        if result is None:
            return jsonify({'error': '文件读取失败'}), 400
        if 'error' in result:
            return jsonify({'error': result['error']}), 400
        return jsonify({
            'originalName': f.filename,
            'months': result['months'],
            'byMonth': result['byMonth'],
        })
    finally:
        try:
            if stored_abs_path.is_file():
                stored_abs_path.unlink()
        except Exception:
            pass


def analyze_sheet_data(path, sheet_name=None):
    """对指定工作表做简单分析：数值列统计、前几行预览。"""
    path = BASE_DIR / path
    if not path.is_file():
        return None
    try:
        df = pd.read_excel(path, sheet_name=sheet_name or 0, header=None)
        # 尝试用第一行作为表头
        if len(df) > 0:
            df_header = pd.read_excel(path, sheet_name=sheet_name or 0, header=0)
            stats = []
            for col in df_header.columns:
                s = df_header[col]
                if pd.api.types.is_numeric_dtype(s):
                    stats.append({
                        'column': str(col),
                        'count': int(s.count()),
                        'mean': round(float(s.mean()), 4) if s.notna().any() else None,
                        'min': round(float(s.min()), 4) if s.notna().any() else None,
                        'max': round(float(s.max()), 4) if s.notna().any() else None,
                        'sum': round(float(s.sum()), 4) if s.notna().any() else None,
                    })
                else:
                    stats.append({
                        'column': str(col),
                        'count': int(s.count()),
                        'unique': int(s.nunique()),
                    })
            preview = df_header.head(20).fillna('').astype(str).values.tolist()
            columns = df_header.columns.astype(str).tolist()
            return {
                'columns': columns,
                'preview': preview,
                'stats': stats,
                'shape': list(df_header.shape),
            }
        return {'preview': [], 'stats': [], 'shape': list(df.shape)}
    except Exception as e:
        return {'error': str(e)}


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/read')
def api_read():
    filename = request.args.get('file')
    if not filename:
        return jsonify({'error': '缺少 file 参数'}), 400
    result = read_excel_sheets(filename)
    if result is None:
        return jsonify({'error': '文件不存在或格式不支持'}), 404
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/analyze')
def api_analyze():
    filename = request.args.get('file')
    sheet = request.args.get('sheet')
    if not filename:
        return jsonify({'error': '缺少 file 参数'}), 400
    result = analyze_sheet_data(filename, sheet)
    if result is None:
        return jsonify({'error': '文件不存在'}), 404
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
